# bchhun, {2020-03-22}

from openpyxl import Workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.utils import get_column_letter

import numpy as np

"""
1- create base template
2- fully parse xml and one well image
3- use output of 1 to call "populate main tab", supplying spotid, proparray, and well name (image file name)
4- repeat 2-3
"""


def create_base_template():
    """
    create an xlsx workbook to mimic that generated by sciReader
    this function generates some basic headers and worksheets

    :return: workbook
    """
    wb = Workbook()

    ws = wb.active
    ws.title = 'main spot report'

    # main report
    ws['A1'].value = 'Test configuration'
    ws['A2'].value = '<NAME OF EXPERIMENT HERE>'

    ws['A6'].value = 'Intensity (median, background corrected) [AU]'
    ws['A7'].value = 'ID'

    ws['B1'].value = 'Configuration version'
    ws['B2'].value = '1.0'

    # REPLICATES TAB
    ws2 = wb.create_sheet(title='Replicates')
    ws2['A1'].value = 'Intensity (median, background corrected) [AU]'
    ws2['A2'].value = 'ID'

    return wb


def populate_main_tab(wb_, spot_id_array_, props_array_, well):
    """
    repeat this function call for EACH well/image that is added to report
        the same workbook is passed to each of those wells

    :param wb_:
    :param spot_id_array_:
    :param props_array_:
    :param well: str
    :return:
    """
    ws = wb_['main spot report']

    # find the next empty column
    maxc = 0
    for c in ws.iter_cols(2, None, 7, 7):
        maxc = c[0].column
    current_coln = maxc+1

    # add well to worksheet
    ws[get_column_letter(current_coln)+str(7)].value = well
    ws[get_column_letter(current_coln)+str(8)].value = well[-1]

    # populating main report Column A --> we do this every well, inefficient
    #   map_spot_cell is dict of "spot-Id" : "Cell position (A1, A2...)"
    map_spot_cell = dict()
    spot_id_array_flat = spot_id_array_.flatten()
    num_spots = len(spot_id_array_flat)
    for r in range(11, 11 + num_spots):
        _ = ws.cell(column=1, row=r, value=spot_id_array_flat[r-11])
        map_spot_cell[spot_id_array_flat[r-11]] = 'A'+str(r)

    # populate main report Well Column
    #   walk through array and check position's cell ID
    #   assign that array value to the xlsx cell based on cell ID
    for row in range(props_array_.shape[0]):
        for col in range(props_array_.shape[1]):
            # get cell corresponding to spot_id, use it to fill next value in xlsx

            if spot_id_array_[row, col] in map_spot_cell:
                pos_letter_idx = column_index_from_string(map_spot_cell[spot_id_array_[row, col]][0]) + current_coln
                pos_num = map_spot_cell[spot_id_array_[row, col]][1]
                # pos should INCREMENT column (letter) and keep row (number) the same
                pos = get_column_letter(pos_letter_idx)+pos_num

                # here we assign mean_intensity but it should be median, background corrected?
                ws[pos].value = props_array_[row, col].mean_intensity

            else:
                raise AttributeError("unable to find cell ID in xlsx columns")

    return wb_


def populate_main_replicates(wb_, props_array_, antigen_array_, well):
    ws = wb_['Replicates']

    # find the next empty column
    maxc = 0
    for c in ws.iter_cols(2, None, 2, 2):
        maxc = c[0].column
    current_coln = maxc + 1

    # add well to worksheet
    ws[get_column_letter(current_coln) + str(2)].value = well
    ws[get_column_letter(current_coln) + str(3)].value = well[-1]

    # populate replicate:  Column A --> we do this every well, inefficient
    #   map_anti_cell is dict of "antigen" : "Cell Position (A1, A2, ...)"
    map_anti_cell = dict()
    antigen_array_flat = antigen_array_.flatten()
    num_spots = len(antigen_array_flat)
    for r in range(6, 6 + num_spots):

        # there are duplicate antigens in the array, must check
        if antigen_array_flat[r-6] in map_anti_cell.keys():
            continue

        _ = ws.cell(column=1, row=r, value=antigen_array_flat[r - 6])
        map_anti_cell[antigen_array_flat[r - 6]] = ['A' + str(r)]

    # populate replicate:  Well Column
    #   there are multiple spots per antigen
    #   need to find them all and then average them. this becomes the cell value

    # create a map from antigen to region_props (multiple)
    map_anti_prop = dict()
    for antigen in map_anti_cell:
        locations = np.where(antigen_array_ == antigen)
        p = [props_array_[coord[0], coord[1]] for coord in list(zip(locations[0], locations[1]))]
        map_anti_prop[antigen] = p

    # iterate through all antigens in Column A
    #   calculate (mean?) value and assign to Well Column
    for antigen, cell in map_anti_cell:
        val = [prop.mean_intensity for prop in map_anti_prop[antigen]]

        pos_letter_idx = column_index_from_string(map_anti_cell[antigen][0]) + current_coln
        pos_num = map_anti_cell[antigen][1]
        pos = get_column_letter(pos_letter_idx)+pos_num

        ws[pos].value = round(np.mean(val), 3)

    return wb_


def populate_cell_tab():
    pass

def populate_cell_replicate():
    pass